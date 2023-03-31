import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json
from datetime import datetime
#Задача: сделать excel-файл с данными о погоде разных столиц мира (типа робота)

#Словарь столиц мира( Сент-Кристофер - исключение)
my_dict = {
"Абу-Даби": "Abu Dhabi",
"Абуджа" : "Abuja",
"Аддис-Абеба" : "Addis Ababa",
"Аккра": "Accra",
"Алжир": "Algeria",
"Амман" : "Amman",
"Амстердам" : "Amsterdam",
"Андорра-ла-Велья":"Andorra la Vella",
"Анкаpа":"Ankara",
"Антананариву":"Antananarivo",
"Апиа":"Apia",
"Асмэра":"Asmara",
"Астана":"Astana",
"Асунсьон":"Asuncion",
"Афины":"Athens",
"Ашхабад":"Ashgabat",
"Багдад":"Baghdad",
"Баку":"Baku",
"Бамако":"Bamako",
"Банги":"Bangui",
"Бангкок":"Bangkok",
"Бандар-Сери-Бегаван":"Bandar Seri Begawan",
"Банжул":"Banjul",
"Сент-Кристофер":"Saint Christopher",
"Бейрут":"Beirut"
}
#берём старый файл в качестве макета, в котором значения будут обновляться 
excel = pd.read_excel('C:/Users/kochkin.ao/Desktop/Weather/Новая_ПогодКА_03.10 08.10.31.xlsx')

# Это api-номер с одного сайта о погоде
api_key = "004729553a7a4cd0a46111936232003"

#Удаляем ненужный столбец
excel.drop(columns=['Unnamed: 0'], axis=1, inplace=True)
print("Одну секунду ...")
# через счётчик перебираем все столицы
i = 0
# Здесь находим температуру через api:

for key in my_dict.keys():
    # Переводим столицы на англ.язык
    q = f'{my_dict[key]}'
    # Вводим в api_url столицы на англ. языке
    api_url = 'http://api.weatherapi.com/v1/current.json?key=004729553a7a4cd0a46111936232003&q={}&aqi=no'.format(q)
    # получаем необработанный текст, т.е набор данных
    response = requests.get(api_url)
    # Проверяем всё ли окей
    if response.status_code == requests.codes.ok:
        # Из непонятного набора информации делаем словарь при помощи json
        data = json.loads(response.text)
        # Обновляем данные
        excel.loc[i, ['Погода']] = 'Нет содержимого' if not data["current"]["temp_c"] else data["current"]["temp_c"]
    else:
        print("Error:", response.status_code, response.text)
        excel.loc[i, ['Погода']] = None
    i += 1

print(excel)
# Создаем счётчик времени, чтоб сохранять данные только в новый файл
current_datetime = datetime.now()
time_now = current_datetime.strftime(" %d.%m.%Y %H.%M.%S")
excel.to_excel(f'C:/Users/kochkin.ao/Desktop/python/RESULT{time_now}.xlsx')   
wb = load_workbook(f'C:/Users/kochkin.ao/Desktop/python/RESULT{time_now}.xlsx')

ws = wb.active
# Находим max значение
max_value = max(excel['Погода'])
# Находим индекс max значения
max_index = excel[excel["Погода"] == max_value ].index.values.astype(str)
# Прибавляем 2, т.к max индекс находится не по встроенному в excel счетчику, а по нарисованному
max_index = f"D{int(max_index) + 2 }"

# Находим min значение
min_value = min(excel['Погода'])
# Находим индекс min значения
min_index = excel[excel["Погода"] == min_value ].index.values.astype(str)
min_index = f"D{int(min_index) + 2 }"
print("Всё хорошо завершено!!!")
#ячейку с max знач. закрашиваем в красный цвет, а с min знач. - в арктический
ws[max_index].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
ws[min_index].fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type = "solid")
# И сохраняем 
wb.save(f'C:/Users/kochkin.ao/Desktop/python/RESULT{time_now}.xlsx')
# Всё!!!